"""Rotas HTTP - PGM Porto Velho - Subprocuradoria Contenciosa."""
from flask import Blueprint, render_template, request, jsonify, current_app
from functools import wraps
import os, json, requests as http
from datetime import datetime, date, timedelta

bp = Blueprint('main', __name__)

# Configurações do Supabase (Chaves originais do seu projeto)
SUPABASE_URL = os.environ.get('SUPABASE_URL', 'https://vmbzykywtgzyxmoxogel.supabase.co')
SUPABASE_KEY = os.environ.get('SUPABASE_KEY', 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InZtYnp5a3l3dGd6eXhtb3hvZ2VsIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc3NjM3NjAyMywiZXhwIjoyMDkxOTUyMDIzfQ.fJlvsVIPLkIg5IXNx1uYfqa5pDj1B8DbRQIUiiTpcEo')

def _sb_headers():
    return {
        'apikey': SUPABASE_KEY,
        'Authorization': f'Bearer {SUPABASE_KEY}',
        'Content-Type': 'application/json',
        'Prefer': 'return=representation'
    }

# --- HELPERS DO SUPABASE ---

def _sb_get(table, params=''):
    try:
        r = http.get(f'{SUPABASE_URL}/rest/v1/{table}?{params}', headers=_sb_headers(), timeout=10)
        return r.json() if r.ok else []
    except: return []

def _sb_post(table, data):
    try:
        r = http.post(f'{SUPABASE_URL}/rest/v1/{table}', headers=_sb_headers(), json=data, timeout=10)
        res = r.json()
        return res[0] if r.ok and isinstance(res, list) and res else res
    except: return {}

def _sb_patch(table, key, val, data):
    try:
        r = http.patch(f'{SUPABASE_URL}/rest/v1/{table}?{key}=eq.{val}', headers=_sb_headers(), json=data, timeout=10)
        return r.ok
    except: return False

def _sb_delete(table, key, val):
    try:
        r = http.delete(f'{SUPABASE_URL}/rest/v1/{table}?{key}=eq.{val}', headers=_sb_headers(), timeout=10)
        return r.ok
    except: return False

# --- GESTÃO DE CACHE ---

def cache_get(chave):
    try:
        rows = _sb_get('dados_cache', f'chave=eq.{chave}&select=valor')
        if rows and isinstance(rows, list) and len(rows) > 0:
            return rows[0].get('valor')
    except: pass
    return None

def cache_set(chave, valor):
    try:
        h = _sb_headers()
        h['Prefer'] = 'resolution=merge-duplicates,return=minimal'
        http.post(f'{SUPABASE_URL}/rest/v1/dados_cache', headers=h,
                  json={'chave': chave, 'valor': valor, 'atualizado': datetime.utcnow().isoformat()}, timeout=10)
    except: pass

# --- LÓGICA DE CÁLCULO DINÂMICO (CORRIGE AS DATAS) ---

def _recalcular_prazos(lista_mestra):
    """Recalcula status e dias restantes baseando-se na data de hoje."""
    hoje = date.today()
    vencidos, proximos, cumpridos_lista = [], [], []
    perf = {}
    manuais = cache_get('cumpridos_manuais') or []

    for p in lista_mestra:
        try:
            dt_prazo = date.fromisoformat(p['data_iso'])
        except: continue
        
        proc = p['processo']
        resp = p['responsavel']
        ja_cumprido = p.get('ja_cumprido', False) or proc in manuais
        
        if resp not in perf: perf[resp] = {'responsavel': resp, 'total': 0, 'cumpridos': 0, 'criticos': 0}
        perf[resp]['total'] += 1

        if ja_cumprido:
            perf[resp]['cumpridos'] += 1
            cumpridos_lista.append(p)
            continue

        diff = (dt_prazo - hoje).days
        p['dias'] = abs(diff)
        
        if diff < 0:
            perf[resp]['criticos'] += 1
            vencidos.append(p)
        elif diff <= 7:
            proximos.append(p)

    perf_list = []
    for r in perf.values():
        r['taxa'] = round(r['cumpridos']/r['total']*100, 1) if r['total'] > 0 else 0
        perf_list.append(r)

    return {
        'stats': {
            'total': len(lista_mestra),
            'vencidos': len(vencidos),
            'proximos': len(proximos),
            'cumpridos': len(cumpridos_lista),
            'taxa': round(len(cumpridos_lista)/len(lista_mestra)*100, 1) if lista_mestra else 0,
            'ultima_atualizacao': hoje.strftime('%d/%m/%Y')
        },
        'performance': sorted(perf_list, key=lambda x: x['taxa'], reverse=True),
        'proximos': sorted(proximos, key=lambda x: x['dias']),
        'vencidos': sorted(vencidos, key=lambda x: x['dias'], reverse=True),
        'cumpridos_lista': cumpridos_lista
    }

def _parse_xlsx(file_obj):
    import openpyxl, warnings
    with warnings.catch_warnings():
        warnings.simplefilter('ignore')
        wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    
    ws = None
    for name in wb.sheetnames:
        if "Prazos" in name or "Pauta" in name:
            ws = wb[name]
            break
    if not ws: ws = wb.active

    lista_completa = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[1] or not row[4]: continue
        prazo_raw = row[1]
        if isinstance(prazo_raw, datetime): prazo_d = prazo_raw.date()
        elif isinstance(prazo_raw, date): prazo_d = prazo_raw
        else: continue
        
        lista_completa.append({
            'processo': str(row[4]).strip(),
            'parte': str(row[5]).strip()[:80] if row[5] else '',
            'responsavel': str(row[2]).strip().upper() if row[2] else 'SEM RESPONSAVEL',
            'prazo': prazo_d.strftime('%d/%m/%Y'),
            'data_iso': prazo_d.isoformat(),
            'vara': str(row[6]).strip() if row[6] else '',
            'ja_cumprido': str(row[13]).strip().upper() in ('SIM', 'PARCIAL', 'PREJUDICADO')
        })
    return lista_completa

# --- AUTH ---

def token_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = request.args.get('token') or request.headers.get('Authorization','').replace('Bearer ','')
        if not token or token != current_app.config['ACCESS_TOKEN']:
            return jsonify({'error':'Token invalido'}), 401
        return f(*args, **kwargs)
    return decorated

# --- ROTAS ---

@bp.route('/api/upload', methods=['POST'])
@token_required
def upload_file():
    file = request.files.get('file')
    if not file: return jsonify({'error':'Arquivo nao fornecido'}), 400
    try:
        lista = _parse_xlsx(file)
        cache_set('lista_mestra', lista)
        res = _recalcular_prazos(lista)
        cache_set('stats', res['stats'])
        cache_set('filename', file.filename)
        return jsonify({'success': True, 'stats': res['stats'], 'filename': file.filename})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@bp.route('/api/dashboard')
@token_required
def get_dashboard():
    lista = cache_get('lista_mestra')
    if not lista: return jsonify({'sem_dados': True})
    return jsonify(_recalcular_prazos(lista))

@bp.route('/api/criticos')
@token_required
def get_criticos():
    lista = cache_get('lista_mestra')
    if not lista: return jsonify({'sem_dados': True})
    res = _recalcular_prazos(lista)
    return jsonify({'vencidos': res['vencidos'], 'proximos': res['proximos']})

@bp.route('/api/equipe', methods=['GET', 'POST'])
@token_required
def gerenciar_equipe():
    if request.method == 'GET':
        membros = _sb_get('equipe', 'order=id.asc')
        return jsonify({'membros': membros})
    data = request.get_json()
    res = _sb_post('equipe', data)
    return jsonify({'success': True, 'membro': res})

@bp.route('/api/equipe/<int:mid>', methods=['DELETE', 'PUT'])
@token_required
def membro_ops(mid):
    if request.method == 'DELETE':
        return jsonify({'success': _sb_delete('equipe', 'id', mid)})
    return jsonify({'success': _sb_patch('equipe', 'id', mid, request.get_json())})

@bp.route('/api/cumprido', methods=['POST'])
@token_required
def marcar_cumprido():
    proc = request.get_json().get('processo')
    if not proc: return jsonify({'error': 'Processo obrigatorio'}), 400
    manuais = cache_get('cumpridos_manuais') or []
    if proc not in manuais: manuais.append(proc)
    cache_set('cumpridos_manuais', manuais)
    return jsonify({'success': True})

@bp.route('/painel')
@token_required
def painel(): return render_template('dashboard.html')

@bp.route('/')
def index():
    from flask import redirect, url_for
    return redirect(url_for('main.painel', token=current_app.config['ACCESS_TOKEN']))
